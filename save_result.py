import json

from openpyxl import Workbook
from openpyxl.styles import Font

from datetime import datetime


class SaveResult:
    def __init__(self, good_dict):
        self.good_dict = good_dict

        self.colums = ['Proposal name', 'Proposed by', 'Date', 'Proposal Text', 'Link', 'Likes', 'Views', 'Category',
                       'Comment Author', 'Comment', 'Comment likes', 'Comment date']

        self.comment_colums = 9


    @staticmethod
    def save_to_json(filename, good_data):
        filename = f'{filename}.json'

        try:
            with open(filename, 'w', encoding='utf-8') as file:
                json.dump(good_data, file, indent=4, ensure_ascii=False)
        except:
            return False

        return filename

    def create_title(self, ws):
        for count, col in enumerate(self.colums):
            ws.cell(row=1, column=count + 1).value = col
            ws.cell(row=1, column=count + 1).font = Font(bold=True)



    def write_data(self, ws, count_def, post):

        ws.cell(row=count_def, column=1).value = post['name_post']
        ws.cell(row=count_def, column=2).value = post['name_author']
        ws.cell(row=count_def, column=3).value = post['date_post']
        ws.cell(row=count_def, column=4).value = post['text_post']
        ws.cell(row=count_def, column=5).value = post['link']
        ws.cell(row=count_def, column=6).value = post['like']
        ws.cell(row=count_def, column=7).value = post['views_post']
        ws.cell(row=count_def, column=8).value = post['name_them']

        if len(post['comments']) == 0:
            return True

        for count_com, comment in enumerate(post['comments']):
            ws.cell(row=count_def + count_com, column=9).value = comment['author_comment']
            ws.cell(row=count_def + count_com, column=10).value = comment['text_comment']
            ws.cell(row=count_def + count_com, column=11).value = int(comment['like_comment'])
            ws.cell(row=count_def + count_com, column=12).value = comment['date_comment']

        return True

    def itter_rows(self, ws):
        count_def = 2
        for count_post, post in enumerate(self.good_dict):
            count_comments = len(post['comments'])


            write_data = self.write_data(ws, count_def, post)

            if count_comments > 1:
                count_def = count_def + count_comments
            else:
                count_def += 1

    def one_sheet(self, ws):

        response = self.create_title(ws)

        response_itter = self.itter_rows(ws)

        return True

    def save_file(self, filename):

        wb = Workbook()

        ws = wb.active

        result = self.one_sheet(ws)

        filename = f'{filename}.xlsx'

        wb.save(filename)

        self.save_to_json(filename, self.good_dict)

        print(f'Сохранил \n{filename}.xlsx\n{filename}.json')

        return filename

